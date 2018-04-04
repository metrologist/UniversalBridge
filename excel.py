from __future__ import division

"""
For interacting with calibration information and bridge readings that have been stored in .xlsx files.
It is derived from ExcelPython.py that was used with CT calibrations.
"""

import math
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import warnings


class CALCULATOR(object):
    def __init__(self, source, output):
        """
        'source' is the spreadsheet with data and 'output' is the name of the
        spreadsheet the results are placed in.
        """
        self.source = source
        self.output = output

    def makeworkbook(self, set, this_title):
        """
        'set' is a list of lists that are each an excel row.
        """
        wb = Workbook()  # create workbook in memory
        ws0 = wb.active  # default active worksheet in position '0'
        ws0.title = this_title
        no_rows = len(set)
        no_columns = len(set[0])
        for i in range(no_rows):
            for j in range(no_columns):
                ws0.cell(row=i + 2, column=j + 1, value=set[i][j])
        try:
            wb.save(filename=self.output)
        except IOError:
            print('Could not create output file in "makeworkbook" method in Excel.py')

    def getdata_block(self, datasheet, block_range):
        """
        'datasheet' is the name of the calc worksheet
        'block_range' is a 4 element list [start row, finish row, start column, finish column]
        """
        warnings.simplefilter('ignore')  # not interested in 'Discarded range with reserved name'
        try:
            wb2 = load_workbook(self.source, data_only=True)  # reads numbers rather than formulae
        except IOError:
            print('Could not open file in "get_data_block" method of Excel.py')
        warnings.simplefilter('default')  # turn warnings back on
        sheet = wb2.get_sheet_by_name(datasheet)
        selected_rows = []
        for i in range(block_range[0], block_range[1] + 1):
            this_row = []
            for j in range(block_range[2], block_range[3] + 1):
                this_row.append(sheet.cell(row=i, column=j).value)
            selected_rows.append(this_row)
        return selected_rows

    def extract_column(self, block, column_number, start_finish):
        """
        Row and column numbers follow R1C1=='A1'
        'block is in format of self.getdata_block
        For extracting specific table columns after having read an entire sheet with get_datablock
        """
        for_output = []
        for i in range(start_finish[0] - 1, start_finish[1]):  # the -1 includes the first row
            for_output.append(block[i][column_number - 1])  # -1 needed as first column is 0
        ##         print for_output
        return for_output


if __name__ == "__main__":
    ianz = CALCULATOR("S21982 1 turn V5.0_draft_KJ.xlsm", "myResults.xlsx")
    # copy a block from the source worksheet
    block_descriptor = [1, 365, 1, 42]
    my_copy_data = ianz.getdata_block('5 A calc', block_descriptor)
    # put the block, unchanged into the output spreadsheet
    ianz.makeworkbook(my_copy_data, 'my_sheet_name')